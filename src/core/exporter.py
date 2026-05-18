from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPORT_TEMPLATE = Template(
    """
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>环境检测核验报告</title>
  <style>
    body { font-family: Arial, 'Microsoft YaHei', sans-serif; margin: 24px; color: #222; }
    h1, h2 { color: #0b5cab; }
    table { width: 100%; border-collapse: collapse; margin-bottom: 16px; }
    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
    th { background: #f4f7fb; }
    .pass { color: #137333; }
    .warn { color: #b06000; }
    .fail { color: #c5221f; }
    .box { background: #f9fbff; border: 1px solid #e5eef8; padding: 12px; margin-bottom: 16px; }
    pre { white-space: pre-wrap; word-break: break-word; }
  </style>
</head>
<body>
  <h1>环境检测智能核验结果</h1>
  <div class="box">
    <strong>文件：</strong>{{ data.file_path }}<br>
    <strong>报告类型：</strong>{{ data.report_type }}<br>
    <strong>报告编号：</strong>{{ data.report_no }}<br>
    <strong>委托单位：</strong>{{ data.commissioning_unit }}<br>
    <strong>受检单位：</strong>{{ data.inspected_unit }}<br>
    <strong>项目名称：</strong>{{ data.project_name }}<br>
    <strong>报告日期：</strong>{{ data.report_date }}<br>
    <strong>CMA：</strong>{{ data.cma_code }} {{ data.cma_valid_from }} ~ {{ data.cma_valid_to }}
  </div>
  <h2>AI / 规则摘要</h2>
  <div class="box"><pre>{{ data.ai_summary }}</pre></div>
  <h2>核验项</h2>
  <table>
    <thead><tr><th>项目</th><th>状态</th><th>说明</th></tr></thead>
    <tbody>
      {% for item in data.checks %}
      <tr>
        <td>{{ item.name }}</td>
        <td class="{{ item.status }}">{{ item.status }}</td>
        <td>{{ item.detail }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  <h2>识别到的检测方法</h2>
  <table>
    <thead><tr><th>项目</th><th>方法/标准号</th></tr></thead>
    <tbody>
      {% for name, method in data.methods_detected.items() %}
      <tr><td>{{ name }}</td><td>{{ method }}</td></tr>
      {% endfor %}
    </tbody>
  </table>
  <h2>原文预览</h2>
  <div class="box"><pre>{{ data.raw_text_preview }}</pre></div>
</body>
</html>
"""
)

SENSOR_TEMPLATE = Template(
    """
<!doctype html>
<html lang="zh-CN">
<head><meta charset="utf-8" /><title>环境监测异常分析</title></head>
<body>
<h1>环境监测异常分析</h1>
<p>{{ data.summary }}</p>
<pre>{{ data | tojson(indent=2, ensure_ascii=False) }}</pre>
</body>
</html>
"""
)

STRUCTURED_TEMPLATE = Template(
    """
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>环境检测报告照片识别结果</title>
  <style>
    body { font-family: Arial, 'Microsoft YaHei', sans-serif; margin: 24px; color: #1f2933; }
    h1, h2 { color: #0b5cab; }
    table { width: 100%; border-collapse: collapse; margin: 12px 0 20px; }
    th, td { border: 1px solid #d7dee8; padding: 7px 8px; text-align: left; vertical-align: top; }
    th { background: #eef4fb; }
    .box { background: #f8fbff; border: 1px solid #dce8f5; padding: 12px; margin-bottom: 16px; }
    .pass { color: #137333; font-weight: 700; }
    .fail { color: #c5221f; font-weight: 700; }
    .unknown, .review { color: #b06000; font-weight: 700; }
    pre { white-space: pre-wrap; word-break: break-word; }
  </style>
</head>
<body>
  <h1>环境检测报告照片识别结果</h1>
  <div class="box">
    <strong>逻辑报告：</strong>{{ data.logical_report_id }}<br>
    <strong>执行标准：</strong>{{ data.standard_name }}<br>
    <strong>增强 PDF：</strong>{{ data.enhanced_pdf_path }}<br>
    <strong>合格率：</strong>{% if data.statistics.judged_count %}{{ "%.1f"|format(data.statistics.compliance_rate * 100) }}%{% else %}暂无可判定数据{% endif %}<br>
    <strong>主要超标污染物：</strong>{{ data.statistics.exceeded_pollutants | join("、") if data.statistics.exceeded_pollutants else "未发现" }}
  </div>
  <h2>简报总结</h2>
  <div class="box"><pre>{{ data.brief_summary }}</pre></div>
  <h2>采样信息</h2>
  <table>
    <thead><tr><th>样点</th><th>采样时间</th><th>频次</th><th>置信度</th><th>来源</th></tr></thead>
    <tbody>
      {% for item in data.sampling %}
      <tr><td>{{ item.sample_point }}</td><td>{{ item.sample_time }}</td><td>{{ item.frequency }}</td><td>{{ item.confidence }}</td><td>{{ item.source_line }}</td></tr>
      {% endfor %}
    </tbody>
  </table>
  <h2>结构化检测记录</h2>
  <table>
    <thead>
      <tr>
        <th>指标</th><th>检测值</th><th>单位</th><th>标准限值</th><th>状态</th><th>数据库匹配</th><th>公式核验</th><th>置信度</th><th>样品/点位</th><th>检测日期</th><th>来源行</th>
      </tr>
    </thead>
    <tbody>
      {% for item in data.records %}
      <tr>
        <td>{{ item.normalized_indicator or item.indicator }}</td>
        <td>{{ item.raw_value or item.value }}</td>
        <td>{{ item.unit }}</td>
        <td>{{ item.standard_limit }} {{ item.limit_unit }}</td>
        <td class="{{ item.status }}">{{ item.status }}{% if item.needs_review %} / 复核{% endif %}</td>
        <td>{{ item.database_match.reason if item.database_match else "" }}</td>
        <td>{{ item.formula_verification.status if item.formula_verification else "" }}</td>
        <td>{{ item.confidence }}</td>
        <td>{{ item.sample_id or item.sample_point }}</td>
        <td>{{ item.detection_date or item.sample_time }}</td>
        <td>{{ item.source_line }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  <h2>可视化建议</h2>
  <ul>
    {% for item in data.visualization_suggestions %}
    <li>{{ item }}</li>
    {% endfor %}
  </ul>
  <h2>预处理页面</h2>
  <table>
    <thead><tr><th>页码</th><th>增强图</th><th>透视置信度</th><th>处理步骤</th></tr></thead>
    <tbody>
      {% for page in data.pages %}
      <tr><td>{{ page.page_index }}</td><td>{{ page.enhanced_image_path }}</td><td>{{ page.transform_confidence }}</td><td>{{ page.preprocessing_steps | join("；") }}</td></tr>
      {% endfor %}
    </tbody>
  </table>
  {% if data.warnings %}
  <h2>警告</h2>
  <div class="box"><pre>{{ data.warnings | join("\n") }}</pre></div>
  {% endif %}
</body>
</html>
"""
)


def export_json(data: Dict[str, Any], output_path: Path) -> Path:
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def export_report_html(data: Dict[str, Any], output_path: Path) -> Path:
    output_path.write_text(REPORT_TEMPLATE.render(data=data), encoding="utf-8")
    return output_path


def export_sensor_html(data: Dict[str, Any], output_path: Path) -> Path:
    output_path.write_text(SENSOR_TEMPLATE.render(data=data), encoding="utf-8")
    return output_path


def export_structured_html(data: Dict[str, Any], output_path: Path) -> Path:
    output_path.write_text(STRUCTURED_TEMPLATE.render(data=data), encoding="utf-8")
    return output_path


def export_structured_excel(data: Dict[str, Any], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "结构化检测记录"
    headers = [
        "指标",
        "检测值",
        "单位",
        "标准限值",
        "限值单位",
        "比较",
        "状态",
        "数据库匹配",
        "公式核验",
        "公式/方法标准",
        "置信度",
        "需复核",
        "样品编号",
        "点位",
        "采样时间",
        "检测日期",
        "频次",
        "来源页",
        "来源行",
        "备注",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    for item in data.get("records", []):
        ws.append(
            [
                item.get("normalized_indicator") or item.get("indicator"),
                item.get("raw_value") or item.get("value"),
                item.get("unit"),
                item.get("standard_limit"),
                item.get("limit_unit"),
                item.get("comparison"),
                item.get("status"),
                (item.get("database_match") or {}).get("reason", ""),
                (item.get("formula_verification") or {}).get("status", ""),
                (item.get("formula_verification") or {}).get("method_standard", ""),
                item.get("confidence"),
                "是" if item.get("needs_review") else "否",
                item.get("sample_id"),
                item.get("sample_point"),
                item.get("sample_time"),
                item.get("detection_date"),
                item.get("frequency"),
                item.get("source_page"),
                item.get("source_line"),
                "；".join(item.get("notes", [])),
            ]
        )

    summary = wb.create_sheet("简报")
    summary_rows = [
        ("逻辑报告", data.get("logical_report_id", "")),
        ("执行标准", data.get("standard_name", "")),
        ("增强 PDF", data.get("enhanced_pdf_path", "")),
        ("合格率", data.get("statistics", {}).get("compliance_rate", 0)),
        ("主要超标污染物", "、".join(data.get("statistics", {}).get("exceeded_pollutants", []))),
        ("简报总结", data.get("brief_summary", "")),
    ]
    for row in summary_rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True)

    review = wb.create_sheet("复核队列")
    review.append(headers)
    for cell in review[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="9C6500")
    for item in data.get("records", []):
        if item.get("needs_review") or item.get("status") in {"unknown", "review"}:
            review.append(
                [
                    item.get("normalized_indicator") or item.get("indicator"),
                    item.get("raw_value") or item.get("value"),
                    item.get("unit"),
                    item.get("standard_limit"),
                    item.get("limit_unit"),
                    item.get("comparison"),
                    item.get("status"),
                    (item.get("database_match") or {}).get("reason", ""),
                    (item.get("formula_verification") or {}).get("status", ""),
                    (item.get("formula_verification") or {}).get("method_standard", ""),
                    item.get("confidence"),
                    "是",
                    item.get("sample_id"),
                    item.get("sample_point"),
                    item.get("sample_time"),
                    item.get("detection_date"),
                    item.get("frequency"),
                    item.get("source_page"),
                    item.get("source_line"),
                    "；".join(item.get("notes", [])),
                ]
            )

    for sheet in wb.worksheets:
        for column in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(length + 2, 10), 50)
    wb.save(output_path)
    return output_path
